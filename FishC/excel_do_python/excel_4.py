#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
four excel program by python
1. styles modules using
2.

"""

import openpyxl


# set cells styles.font
"""
openpyxl.styles.Font
"""

wb = openpyxl.Workbook()

ws = wb.active

b2 = ws['B2']

b2.value = "FishC"

# bold and color is red
bold_red_font = openpyxl.styles.Font(bold=True, color="FF0000")

b2.font = bold_red_font

b3 = ws['B3']
b3.value = "FishC"
italic_strike_blue_16font = openpyxl.styles.Font(size=16, italic=True, strike=True, color="0000FF")
b3.font = italic_strike_blue_16font

# fill cells
yellow_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FFFF00")
b2.fill = yellow_fill

red2green_fill = openpyxl.styles.GradientFill(type="linear", stop=("FF0000", "00FF00"))
b3.fill = red2green_fill

# set Border and Side
thin_side = openpyxl.styles.Side(border_style="thin", color="000000")

double_side = openpyxl.styles.Side(border_style="double", color="0000FF")

b2.border = openpyxl.styles.Border(diagonal=thin_side, diagonalUp=True, diagonalDown=True)
b3.border = openpyxl.styles.Border(left=double_side, top=double_side, right=double_side, bottom=double_side)

# text alignment

ws.merge_cells('A1:C2')
ws['A1'].value = "I love FishC.com"

center_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
ws['A1'].alignment = center_alignment

# name styles
hightlight = openpyxl.styles.NamedStyle(name="highlight")
hightlight.font = openpyxl.styles.Font(bold=True, size=20)
hightlight.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

wb.add_named_style(hightlight)
ws['A1'].style = hightlight
ws['B5'].value = "Love"
ws['B5'].style = hightlight

wb.save("font_set.xlsx")


