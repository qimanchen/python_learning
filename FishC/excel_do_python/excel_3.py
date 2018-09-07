#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
third excel program by python
1. personal work label
2.

"""


import openpyxl

# create a work book
wb = openpyxl.Workbook()

# create work sheet
ws1 = wb.create_sheet(title="Test1")
ws2 = wb.create_sheet(title="Test2")
ws3 = wb.create_sheet(title="Test3")
ws4 = wb.create_sheet(title="Test4")

# worksheet tab set
# RGB set
ws1.sheet_properties.tabColor = "FF0000"  # red
ws2.sheet_properties.tabColor = "00FF00"  # greed
ws3.sheet_properties.tabColor = "0000FF"  # blue
ws4.sheet_properties.tabColor = "8B008B"  # special color - purple

# set row height and column width
"""
row:
    row_dimensions[2].height

column:
    column_dimensions['C'].width
    
Note!
    row and column's unit is not same.
"""
ws2.row_dimensions[2].height = 100
ws2.column_dimensions['C'].width = 50

# cell merge and unmerge
"""
merge_cells("A1:C3")
Example:
    A   B   C
  1
  2
  3
  
unmerge_cells()

"""
# merge cells

ws1.merge_cells("A1:C3")

ws1['A1'] = "I love FishC.com!"

# unmerge cells
ws1.unmerge_cells("A1:C3")
# Note! --- must be same area with merge

# freeze window
# set
ws1.freeze_panes = "B8"

# un freeze
ws1.freeze_panes = None
# or ws1.freeze_panes = "A1"


# save workbook
wb.save("tab_set.xlsx")

