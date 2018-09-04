#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
first excel program by python
creat and save excel file

set openpyxl
pip  install openpyxl
"""

import openpyxl
import datetime

# create a work book
wb = openpyxl.Workbook()

# get work table that is working
ws = wb.active

# print work table name
print("table", ws.title)

# add date to work table
ws['A1'] = 520  # raw - one, column - one
for i in range(2, 5):
    print('A' + str(i))
    ws['A'+str(i)] = i**i

ws.append([1, 2, 3])  # raw - two, column - one to three

# add time date from datetime module
ws['A3'] = datetime.datetime.now()  # real time

# save work book and name to 'first_demo.xlsx'
wb.save("first_demo.xlsx")






