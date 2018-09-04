#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
second excel program by python
1. open excel file
2.

"""


import openpyxl

# open a excel file -> work book
wb = openpyxl.load_workbook("first_demo.xlsx")
# openpyxl.load_workbook(r"H:\gitHub_project\python_learning\FishC\excel_do_python\first_demo.xlsx")

# type(wb) --> check wb type

# get work table
# print(wb.get_sheet_names())
# get_sheet_names() replaced by wb.sheetnames

print(wb.sheetnames)

# get special name work table
# if input table name is not exist that is raise error
ws = wb['Sheet']
# get_sheet_by_name('Sheet') replaced by wb[sheetname]

# create work sheet
# index  --  create position
# title  --  work table name
nws = wb.create_sheet(index=0, title="FishC DEMO")

# print all work table names
print(wb.sheetnames)

# delete a work sheet --> must input a work table object
del wb["FishC DEMO"]
# remove_sheet replaced by wb.remove(worksheet) or del wb[sheetname]

# local any cell
c = ws['A2']
print(c.row)
print(c.column)
print(c.coordinate)
print(ws['A2'].value)

# get column number 496 -- 'SB'
print(openpyxl.cell.cell.get_column_letter(496))
# 'JB' -- 262 column
print(openpyxl.cell.cell.column_index_from_string('JB'))

# move cell
d = c.offset(2, 2)
print(d.value)

# find more cell
for each_movie in ws['A2': 'B10']:
    for each_cell in each_movie:
        print(each_cell.value, end=' ')
    print('\n')

# ws.rows all row
for each_row in ws.rows:
    print(each_row[0].value)

for each_row in ws.iter_rows(min_row=2, min_col=1, max_row=4, max_col=2):
    print(each_row[0].value)

# copy work sheet
new = wb.copy_worksheet(ws)
type(new)
wb.save("first_demo.xlsx")





