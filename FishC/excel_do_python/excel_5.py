#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
five excel program by python
1. number styles
2.

"""

import openpyxl
from openpyxl.styles.colors import RED, GREEN, BLUE
import datetime


wb = openpyxl.Workbook()
ws = wb.active

# set number styles
ws.append(['type', 'text', "number"])
# add text and number date
ws['B2'] = '520'
ws['C2'] = 520

ws2 = wb.create_sheet(title="set_number_date_styles")

# set number date formate
ws2['A1'] = 88.8
ws2['A1'].number_format = "#,###.00鱼币"

ws2['A2'] = datetime.datetime.today()
ws2['A2'].number_format = "yyyy - mm - dd"

# set format
ws3 = wb.create_sheet(title="set_number")
# set more than one date format
# 正值；负值；零值；文本

ws3['A1'].number_format = "[RED]+#,###.00;[GREEN]-#,###.00"
ws3['A1'] = 99

ws3['A2'].number_format = "[RED]+#,###.00;[GREEN]-#,###.00"
ws3['A2'] = -99

ws3['A3'].number_format = "[RED];[GREEN];[BLUE];[YELLOW]"
ws3['A3'] = 0

ws3['A4'].number_format = "[RED];[GREEN];[BLUE];[YELLOW]"
ws3['A4'] = "FishC"

# set additional conditions
# Note! the string must be Chinese
ws3['A5'].number_format = "[=1]男;[=0]女"
ws3['A5'] = 0

ws3['A6'].number_format = "[=1]男;[=0]女"
ws3['A6'] = 1

ws3['A7'].number_format = "[=1]男;[=0]女"
ws3['A7'] = 0

# set score greed judge
ws3['A8'].number_format = "[<60][RED]不及格;[>=60][GREEN]及格"
ws3['A8'] = 58

ws3['A9'].number_format = "[<60][RED]不及格;[>=60][GREEN]及格"
ws3['A9'] = 68

wb.save("date_styles.xlsx")

